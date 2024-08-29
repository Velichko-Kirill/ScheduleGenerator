import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from GeneticAlgorithm import Config
from pathlib import Path

class ScheduleExporter:
    def __init__(self, best_individual, encoding, data, file_name=Path().resolve() / "data" / "output_schedule.xlsx"):
        self.best_individual = best_individual
        self.encoding = encoding
        self.file_name = file_name
        self.config = Config()
        self.data = data

    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"

        # Создание заголовков групп (строка 1)
        ws.cell(row=1, column=1, value="День/Слоты")
        for col in range(self.config.n_groups):
            ws.cell(row=1, column=col + 2, value=f"Группа {col+1}")  # Starting from column 2

        # Установка ширины столбцов и форматирования
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20  # Increased width

        # Заполнение расписания
        current_row = 2
        for day in range(self.config.n_days):
            day_name = f"День {day + 1}"
            ws.cell(row=current_row, column=1, value=day_name)
            current_row += 1

            for slot in range(self.config.n_hours):
                ws.cell(row=current_row, column=1, value=f"Слот {slot + 1}")

                for group_idx in range(self.config.n_groups):
                    idx = day * self.config.n_hours + slot + group_idx * self.config.n_days * self.config.n_hours

                    try:
                        triple_idx = self.best_individual[idx]
                        teacher_idx, subject_idx, classroom_idx = self.encoding[triple_idx]

                        # Get actual teacher, subject, and classroom names/numbers
                        teacher_name = list(self.data.teachers.keys())[teacher_idx]

                        # Assuming subjects are the same for all groups (modify if needed)
                        subject_name = list(list(self.data.groups.values())[0].subjects.keys())[subject_idx] 
                        classroom_number = self.data.rooms[classroom_idx].number

                        cell_value = f"Учитель: {teacher_name}\n" \
                                      f"Предмет: {subject_name}\n" \
                                      f"Аудитория: {classroom_number}"

                        cell = ws.cell(row=current_row, column=group_idx + 2, value=cell_value)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    except IndexError:
                        cell = ws.cell(row=current_row, column=group_idx + 2, value="N/A")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                current_row += 1

            current_row += 1  # Add an empty row between days

        # Автоподбор высоты строк
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = cell.value.split('\n')
                    height = len(lines) * 12  # Adjust 12 for font size
                    if height > max_height:
                        max_height = height
            ws.row_dimensions[cell.row].height = max_height if max_height > 0 else 20  # Minimum height

        # Сохранение файла
        wb.save(self.file_name)
        print(f"Расписание успешно экспортировано в файл {self.file_name}")