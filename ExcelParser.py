from openpyxl import load_workbook
from Entities import *

class ExcelParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename=self.file_path, data_only=True)
        self.data = MainData()  # Главный объект, содержащий все данные

    def setup(self):
        self.parse_time_restrictions()
        self.parse_rooms()
        self.parse_curriculum()
        self.parse_equipment_requirements()
        self.parse_teacher_matrix()
        self.parse_day_and_time_constraints()

    def get_sheet(self, sheet_name):
        return self.workbook[sheet_name]

    def parse_rooms(self):
        sheet = self.get_sheet('Аудитории')

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            room_number = row[0]  # Номер аудитории (столбец A)

            if not room_number:
                continue

            building_flag = row[1]  # Другой корпус (столбец B)
            equipment = row[2]  # Оборудование (столбец C)

            building = "Другой корпус" if building_flag == "Да" else "Главный корпус"
            equipment = equipment if equipment != "Нет" else None

            # Создаем объект Room и добавляем его в MainData
            room = Room(number=room_number, building=building, equipment=equipment, room_type=None)
            self.data.add_room(room)

    def parse_day_and_time_constraints(self):
        sheet = self.get_sheet('Ограничения по дням и времени')

        # Обрабатываем каждую строку с 4-й строки
        for row in sheet.iter_rows(min_row=4, values_only=True):
            subject_name = row[0]  # Название предмета (столбец A)

            if not subject_name:
                continue

            # Обработка ограничений по дням недели (столбцы B-G)
            for day_index in range(6):  # Индексы для дней недели: 0 = понедельник, 5 = суббота
                day_constraint = row[day_index + 1]  # Смещение на +1 из-за столбца A

                if day_constraint == 'X':
                    # Применяем ограничения по дням недели для каждого предмета в каждой группе
                    for group in self.data.groups.values():
                        if subject_name in group.subjects:
                            group.subjects[subject_name].day_constraints[day_index] = True

            # Обработка ограничений по временным слотам (столбцы H-O)
            for slot_index in range(8):  # Индексы для слотов: 0 = 1-ое занятие, 7 = 8-ое занятие
                time_constraint = row[slot_index + 7]  # Смещение на +7 из-за столбцов A-G

                if time_constraint == 'X':
                    # Применяем ограничения по временным слотам для каждого предмета в каждой группе
                    for group in self.data.groups.values():
                        if subject_name in group.subjects:
                            group.subjects[subject_name].time_constraints[slot_index] = True

    def parse_curriculum(self):
        sheet = self.get_sheet('Учебный план')

        group_columns = ["C", "D", "E", "F", "G", "I", "J", "K", "L", "M"]

        for row in range(4, sheet.max_row + 1):
            subject_name = sheet[f"A{row}"].value

            if not subject_name:
                continue

            for col in group_columns:
                cell = sheet[f"{col}{row}"]
                group_name = sheet[f"{col}3"].value

                if group_name and cell.value is not None:
                    hours = cell.value
                    is_difficult = cell.fill.start_color.index != "00000000" and cell.fill.start_color.index != "000000"

                    if group_name not in self.data.groups:
                        group = ClassGroup(group_name)
                        self.data.add_group(group)
                    else:
                        group = self.data.groups[group_name]

                    # Если предмет не существует в группе, добавляем его
                    if subject_name not in group.subjects:
                        subject_info = Subject(subject_name, weekly_hours=hours, is_difficult=is_difficult)
                        group.add_subject(subject_info)
                    else:
                        group.subjects[subject_name].weekly_hours = hours
                        group.subjects[subject_name].is_difficult = is_difficult

    def parse_time_restrictions(self):
        sheet = self.get_sheet('Ограничения по часам')

        daily_hours_limit = sheet['B1'].value
        weekly_hours_limit = sheet['B2'].value
        daily_difficult_hours_limit = sheet['B3'].value

        schedule_config = ScheduleConfig(daily_hours_limit, weekly_hours_limit, daily_difficult_hours_limit)
        self.data.set_schedule_config(schedule_config)

    def parse_equipment_requirements(self):
        sheet = self.get_sheet('Требования по оснащению')

        group_columns = ["C", "D", "E", "F", "G", "I", "J", "K", "L", "M"]

        for row in range(4, sheet.max_row + 1):
            subject_name = sheet[f"A{row}"].value  # Название предмета (столбец A)

            if not subject_name:
                continue

            for col in group_columns:
                cell = sheet[f"{col}{row}"]
                group_name = sheet[f"{col}3"].value  # Название группы в строке 3

                if group_name and cell.value:
                    equipment_requirement = cell.value

                    # Проверка на знак "-", если есть, заменяем на None
                    if equipment_requirement == "-":
                        equipment_requirement = None

                    if group_name not in self.data.groups:
                        group = ClassGroup(group_name)
                        self.data.add_group(group)
                    else:
                        group = self.data.groups[group_name]

                    if subject_name in group.subjects:
                        group.subjects[subject_name].equipment_requirement = equipment_requirement
                    else:
                        group_subject = Subject(subject_name, equipment_requirement=equipment_requirement)
                        group.add_subject(group_subject)

    def parse_teacher_matrix(self):
        sheet = self.get_sheet('Матрицы преподавателей')

        group_columns = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        groups = [sheet[f"{col}2"].value for col in group_columns]

        for row in range(3, sheet.max_row + 1):
            teacher_name = sheet[f"A{row}"].value  # Имя преподавателя (столбец A)
            subjects_string = sheet[f"B{row}"].value  # Предметы, которые ведет преподаватель (столбец B)

            if not teacher_name or not subjects_string:
                continue

            subjects = [subject.strip() for subject in subjects_string.split(",")]

            if teacher_name not in self.data.teachers:
                teacher = Teacher(teacher_name)
                self.data.add_teacher(teacher)
            else:
                teacher = self.data.teachers[teacher_name]

            teacher.subjects_name.extend(subjects)

            for col_index, group_name in enumerate(groups):
                cell = sheet[f"{group_columns[col_index]}{row}"]

                if cell.value == "X":
                    if group_name not in teacher.available_groups_name:
                        teacher.available_groups_name.append(group_name)

                    if group_name in self.data.groups:
                        group = self.data.groups[group_name]
                    else:
                        group = ClassGroup(group_name)
                        self.data.add_group(group)

                    for subject_name in subjects:
                        if subject_name not in group.subjects:
                            group_subject = Subject(subject_name)
                            group.add_subject(group_subject)
